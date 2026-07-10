"""
Fighter's OS — Excel Builder
Run this Python script to generate fighters-os.xlsx
Then upload to Google Drive and open in Google Sheets.

Usage: python build_sheet.py
"""

import subprocess, sys

def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    install("openpyxl")
    import openpyxl

import csv, os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# COLOURS
# ─────────────────────────────────────────────
def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col.lstrip("#"))

def font(hex_col, bold=False, size=10, name="Calibri"):
    return Font(color=hex_col.lstrip("#"), bold=bold, size=size, name=name)

BG      = "0a0a14"
PANEL   = "13132a"
INPUT   = "1e1e3f"
HEADER  = "110000"
DIVIDER = "1e1e3a"
RED     = "e63946"
GOLD    = "ffd60a"
GREEN   = "06d6a0"
AMBER   = "ffb703"
BLUE    = "48cae4"
WHITE   = "ffffff"
DIM     = "8888aa"
LABEL   = "a0a8cc"
BODY    = "d0d0e8"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def style(ws, row, col, value="", bg=PANEL, fg=BODY, bold=False,
          size=10, halign="left", valign="center", wrap=False, formula=False):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.fill = fill(bg)
    c.font = font(fg, bold=bold, size=size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    return c

def merge_style(ws, row, col_start, col_end, value="", bg=PANEL, fg=BODY,
                bold=False, size=10, halign="left", valign="center", wrap=False):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value = value
    c.fill = fill(bg)
    c.font = font(fg, bold=bold, size=size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    return c

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def divider_row(ws, row, num_cols=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    ws.cell(row=row, column=1).fill = fill(DIVIDER)
    row_height(ws, row, 4)

def block_header(ws, row, text, bg=RED, num_cols=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.fill = fill(bg)
    c.font = font(WHITE, bold=True, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    row_height(ws, row, 26)

# Lookup formula builder — uses simple sheet names (no emoji, no quotes needed)
# Format: "P"&Settings!B3&"-D"&MID(B4,5,1)&"-BLOCK-SLOT-VARIANT"
def lookup(field_col, key_suffix, playbook="Playbook", settings="Settings"):
    # field_col is the column letter in the Playbook sheet
    # key_suffix e.g. "-MOB-1-STD"
    key = f'"P"&{settings}!B3&"-D"&MID(B4,5,1)&"{key_suffix}"'
    return f'=IFERROR(INDEX({playbook}!{field_col}:{field_col},MATCH({key},{playbook}!A:A,0)),"")'

def lookup_ha(field_col, slot, block, playbook="Playbook", settings="Settings"):
    """Hip-aware lookup: if H4<=2 and HA variant exists, use HA; else STD"""
    key_std = f'"P"&{settings}!B3&"-D"&MID(B4,5,1)&"-{block}-{slot}-STD"'
    key_ha  = f'"P"&{settings}!B3&"-D"&MID(B4,5,1)&"-{block}-{slot}-HA"'
    std = f'IFERROR(INDEX({playbook}!{field_col}:{field_col},MATCH({key_std},{playbook}!A:A,0)),"")'
    ha  = f'IFERROR(INDEX({playbook}!{field_col}:{field_col},MATCH({key_ha},{playbook}!A:A,0)),"")'
    return f'=IF(H4<=2,IF({ha}<>"","🔴 "&{ha},{std}),{std})'

# ─────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────
wb = Workbook()

# ── SHEET ORDER ──────────────────────────────
# Remove default sheet, create in display order
wb.remove(wb.active)

ws_hud  = wb.create_sheet("HUD")
ws_set  = wb.create_sheet("Settings")
ws_play = wb.create_sheet("Playbook")
ws_log  = wb.create_sheet("FightLog")
ws_stat = wb.create_sheet("Stats")

# Tab colours
ws_hud.sheet_properties.tabColor  = RED
ws_set.sheet_properties.tabColor  = "2a2a4a"
ws_play.sheet_properties.tabColor = "3d3d5c"
ws_log.sheet_properties.tabColor  = "1a3a1a"
ws_stat.sheet_properties.tabColor = "1a1a3a"

# ─────────────────────────────────────────────
# PLAYBOOK SHEET
# ─────────────────────────────────────────────
print("Building Playbook...")

pb_headers = [
    "Key","Phase","Day","Block","Slot","Variant",
    "Exercise","Sets","Target_Reps","Load_Note",
    "PAP_Exercise","PAP_Sets","PAP_Reps",
    "Combo_Focus","Cue"
]
for ci, h in enumerate(pb_headers, 1):
    c = ws_play.cell(row=1, column=ci, value=h)
    c.fill = fill("1a1a2e")
    c.font = font(WHITE, bold=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Load CSV
script_dir = os.path.dirname(os.path.abspath(__file__))
csv_path = os.path.join(script_dir, "playbook.csv")
pb_rows = []
with open(csv_path, newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    next(reader)  # skip header (we wrote it above)
    for row in reader:
        if row:
            pb_rows.append(row)

for ri, row in enumerate(pb_rows, 2):
    for ci, val in enumerate(row[:15], 1):
        c = ws_play.cell(row=ri, column=ci, value=val.strip())
        c.fill = fill(PANEL if ri % 2 == 0 else "0f0f20")
        c.font = font(BODY, size=9)
        c.alignment = Alignment(vertical="center", wrap_text=True)

col_width(ws_play, 1, 28)
col_width(ws_play, 7, 35)
col_width(ws_play, 14, 40)
col_width(ws_play, 15, 40)
ws_play.freeze_panes = "A2"
ws_play.sheet_state = "hidden"
print(f"  → {len(pb_rows)} exercise rows loaded")

# ─────────────────────────────────────────────
# SETTINGS SHEET
# ─────────────────────────────────────────────
print("Building Settings...")

merge_style(ws_set, 1, 1, 3, "⚙️  FIGHTER'S OS — SETTINGS",
            bg=RED, fg=WHITE, bold=True, size=14, halign="center")
row_height(ws_set, 1, 40)

settings_rows = [
    (2,  "",                 "",  ""),
    (3,  "CURRENT PHASE",    1,   "← Change to 2 or 3 when unlocked"),
    (4,  "SESSIONS PER PHASE (threshold)", 12, "Phase 2 unlocks after 12 completed gym sessions"),
    (5,  "",                 "",  ""),
    (6,  "PHASE 1 GYM SESSIONS LOGGED",
         "=COUNTIFS(FightLog!C:C,1,FightLog!B:B,\"<>2\",FightLog!B:B,\"<>4\")",
         "Auto-calculated"),
    (7,  "PHASE 2 GYM SESSIONS LOGGED",
         "=COUNTIFS(FightLog!C:C,2,FightLog!B:B,\"<>2\",FightLog!B:B,\"<>4\")",
         "Auto-calculated"),
    (8,  "PHASE 3 GYM SESSIONS LOGGED",
         "=COUNTIFS(FightLog!C:C,3,FightLog!B:B,\"<>2\",FightLog!B:B,\"<>4\")",
         "Auto-calculated"),
    (9,  "",                 "",  ""),
    (10, "PHASE UNLOCK STATUS",
         "=IF(B6>=B4,\"PHASE \"&(B3+1)&\" UNLOCKED — Update phase above!\",\"Sessions until Phase \"&(B3+1)&\": \"&(B4-B6))",
         "Auto-updates"),
    (11, "",                 "",  ""),
    (12, "ATHLETE",          "Fighter's OS User",  "Edit your name"),
    (13, "BRANCH",           "FlyeFit Macken St.", ""),
]

for row_n, label, val, note in settings_rows:
    style(ws_set, row_n, 1, label, bg=PANEL, fg=LABEL, bold=bool(label))
    c = ws_set.cell(row=row_n, column=2, value=val)
    c.fill = fill(INPUT)
    c.font = font(GOLD if row_n == 3 else WHITE, bold=(row_n==3), size=14 if row_n==3 else 10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    style(ws_set, row_n, 3, note, bg=PANEL, fg=DIM, size=9)

col_width(ws_set, 1, 38)
col_width(ws_set, 2, 22)
col_width(ws_set, 3, 40)

# ─────────────────────────────────────────────
# HUD SHEET
# ─────────────────────────────────────────────
print("Building HUD...")

# Column widths
col_width(ws_hud, 1, 34)  # A: Labels
col_width(ws_hud, 2, 11)  # B
col_width(ws_hud, 3, 9)   # C
col_width(ws_hud, 4, 11)  # D
col_width(ws_hud, 5, 9)   # E
col_width(ws_hud, 6, 11)  # F
col_width(ws_hud, 7, 9)   # G
col_width(ws_hud, 8, 11)  # H
col_width(ws_hud, 9, 9)   # I

# Fill all cells with base BG first
for r in range(1, 65):
    for c in range(1, 10):
        ws_hud.cell(r, c).fill = fill(BG)

# ── ROW 1: Title ─────────────────────────────
merge_style(ws_hud, 1, 1, 9, "⚔️  F I G H T E R ' S   O S",
            bg=RED, fg=WHITE, bold=True, size=22, halign="center")
row_height(ws_hud, 1, 55)

# ── ROW 2: Subtitle ──────────────────────────
merge_style(ws_hud, 2, 1, 9, "FlyeFit Macken Street  |  Combat Performance System",
            bg=HEADER, fg=DIM, size=9, halign="center")
row_height(ws_hud, 2, 22)

# ── ROW 3: Divider ───────────────────────────
divider_row(ws_hud, 3)

# ── ROW 4: Selector row ──────────────────────
style(ws_hud, 4, 1, "SELECT YOUR DAY", bg=PANEL, fg=LABEL, bold=True)
style(ws_hud, 4, 2, "Day 1", bg=INPUT, fg=GOLD, bold=True, halign="center")  # dropdown target
style(ws_hud, 4, 3, "", bg=PANEL)
style(ws_hud, 4, 4, "PHASE", bg=PANEL, fg=LABEL, bold=True)
style(ws_hud, 4, 5, "=\"Phase \"&Settings!B3", bg=INPUT, fg=WHITE, bold=True, halign="center")
style(ws_hud, 4, 6, "HIP SCORE (1-5)", bg=PANEL, fg=LABEL, bold=True)
style(ws_hud, 4, 7, 3, bg=INPUT, fg=WHITE, bold=True, halign="center")  # dropdown target
style(ws_hud, 4, 8, "=IF(G4<=2,\"🔴 HIGH ALERT\",IF(G4=3,\"🟡 MODERATE\",\"🟢 GOOD\"))",
      bg=PANEL, fg=WHITE, bold=True, halign="center")
ws_hud.merge_cells("H4:I4")
row_height(ws_hud, 4, 32)

# ── ROW 5: Fight gym notice ───────────────────
merge_style(ws_hud, 5, 1, 9,
    "=IF(OR(B4=\"Day 2\",B4=\"Day 4\"),\"🥊  FIGHT GYM DAY — Log your session in the Notes section below\",\"\")",
    bg=PANEL, fg=GREEN, bold=True, size=11, halign="center")
row_height(ws_hud, 5, 26)

# ── ROW 6: Divider ───────────────────────────
divider_row(ws_hud, 6)

# ── ROW 7: MOBILITY HEADER ───────────────────
block_header(ws_hud, 7, "🔥  MOBILITY  &  INJURY PREP", bg=AMBER)

# ── ROW 8: Mob column headers ─────────────────
style(ws_hud, 8, 1, "Movement", bg=PANEL, fg=LABEL, bold=True, size=9)
style(ws_hud, 8, 2, "✓ Done", bg=PANEL, fg=LABEL, bold=True, size=9, halign="center")
style(ws_hud, 8, 3, "Duration", bg=PANEL, fg=LABEL, bold=True, size=9)
ws_hud.merge_cells("D8:I8")
style(ws_hud, 8, 4, "Coach Cue", bg=PANEL, fg=LABEL, bold=True, size=9)
row_height(ws_hud, 8, 22)

# ── ROWS 9-13: Mobility slots ─────────────────
for slot in range(1, 6):
    r = 8 + slot
    # Exercise name (hip-aware)
    ws_hud.cell(r, 1).value = lookup_ha("G", slot, "MOB")
    ws_hud.cell(r, 1).fill  = fill(PANEL)
    ws_hud.cell(r, 1).font  = font(BODY, size=10)
    ws_hud.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
    # Checkbox placeholder (script will add real checkboxes)
    style(ws_hud, r, 2, False, bg=INPUT, halign="center")
    # Duration
    ws_hud.cell(r, 3).value = lookup("I", f"-MOB-{slot}-STD")
    ws_hud.cell(r, 3).fill  = fill(PANEL)
    ws_hud.cell(r, 3).font  = font(DIM, size=9)
    ws_hud.cell(r, 3).alignment = Alignment(vertical="center")
    # Cue (merged D-I)
    ws_hud.merge_cells(f"D{r}:I{r}")
    ws_hud.cell(r, 4).value = lookup_ha("O", slot, "MOB")
    ws_hud.cell(r, 4).fill  = fill(PANEL)
    ws_hud.cell(r, 4).font  = font(DIM, size=9)
    ws_hud.cell(r, 4).alignment = Alignment(vertical="center", wrap_text=True)
    row_height(ws_hud, r, 30)

# ── ROW 14: Divider ───────────────────────────
divider_row(ws_hud, 14)

# ── ROW 15: STRENGTH HEADER ───────────────────
block_header(ws_hud, 15, "💪  STRENGTH  +  PAP  SUPERSETS", bg=RED)

# ── ROW 16: Set column headers ─────────────────
set_labels = ["Exercise", "S1 kg", "S1 reps", "S2 kg", "S2 reps",
              "S3 kg", "S3 reps", "S4 kg", "S4 reps"]
for ci, lbl in enumerate(set_labels, 1):
    style(ws_hud, 16, ci, lbl, bg=PANEL, fg=LABEL, bold=True, size=9, halign="center")
row_height(ws_hud, 16, 22)

# ── ROWS 17-40: 4 exercises × 6 rows each ─────
# Structure per exercise:
#   row+0: exercise name + target (merged cols B-H), cue col I
#   row+1..4: Set 1-4 | load | reps | load | reps | load | reps | load | reps
#   row+5: PAP row
str_start = 17
for ex in range(1, 5):
    er = str_start + (ex - 1) * 6  # exercise start row
    str_key    = f"-STR-{ex}"
    str_key_b  = f'"P"&Settings!B3&"-D"&MID(B4,5,1)&"{str_key}"'

    # Exercise name row
    ex_name_f = f'=IF(IFERROR(INDEX(Playbook!G:G,MATCH({str_key_b},Playbook!A:A,0)),"")="","","[{chr(64+ex)}] "&IFERROR(INDEX(Playbook!G:G,MATCH({str_key_b},Playbook!A:A,0)),""))'
    ws_hud.cell(er, 1).value = ex_name_f
    ws_hud.cell(er, 1).fill  = fill(INPUT)
    ws_hud.cell(er, 1).font  = font(WHITE, bold=True, size=10)
    ws_hud.cell(er, 1).alignment = Alignment(vertical="center", wrap_text=True)

    tgt_f = f'=IFERROR("Target: "&INDEX(Playbook!I:I,MATCH({str_key_b},Playbook!A:A,0))&"  |  "&INDEX(Playbook!J:J,MATCH({str_key_b},Playbook!A:A,0)),"")'
    ws_hud.merge_cells(f"B{er}:H{er}")
    ws_hud.cell(er, 2).value = tgt_f
    ws_hud.cell(er, 2).fill  = fill(INPUT)
    ws_hud.cell(er, 2).font  = font(DIM, size=9)
    ws_hud.cell(er, 2).alignment = Alignment(vertical="center")

    cue_f = f'=IFERROR("💡 "&INDEX(Playbook!O:O,MATCH({str_key_b},Playbook!A:A,0)),"")'
    ws_hud.cell(er, 9).value = cue_f
    ws_hud.cell(er, 9).fill  = fill(INPUT)
    ws_hud.cell(er, 9).font  = font(AMBER, size=8)
    ws_hud.cell(er, 9).alignment = Alignment(vertical="center", wrap_text=True)
    row_height(ws_hud, er, 28)

    # Set rows
    for s in range(1, 5):
        sr = er + s
        style(ws_hud, sr, 1, f"Set {s}", bg=PANEL, fg=LABEL, size=9, halign="right")
        for col in range(2, 10):
            style(ws_hud, sr, col, "", bg=INPUT, fg=WHITE, halign="center")
        row_height(ws_hud, sr, 26)

    # PAP row
    pr = er + 5
    pap_f = (f'=IFERROR(IF(INDEX(Playbook!K:K,MATCH({str_key_b},Playbook!A:A,0))="",'
             f'"","⚡ PAP: "&INDEX(Playbook!K:K,MATCH({str_key_b},Playbook!A:A,0))'
             f'&"  ("&INDEX(Playbook!L:L,MATCH({str_key_b},Playbook!A:A,0))'
             f'&"×"&INDEX(Playbook!M:M,MATCH({str_key_b},Playbook!A:A,0))&")")"")')
    # Fix the formula string (close properly)
    pap_f = (f'=IFERROR("⚡ PAP: "&INDEX(Playbook!K:K,MATCH({str_key_b},Playbook!A:A,0))'
             f'&"  ("&INDEX(Playbook!L:L,MATCH({str_key_b},Playbook!A:A,0))'
             f'&"×"&INDEX(Playbook!M:M,MATCH({str_key_b},Playbook!A:A,0))&")","—")')
    ws_hud.cell(pr, 1).value = pap_f
    ws_hud.cell(pr, 1).fill  = fill("1a0a00")
    ws_hud.cell(pr, 1).font  = font(AMBER, bold=True, size=10)
    ws_hud.cell(pr, 1).alignment = Alignment(vertical="center")
    # PAP checkbox col
    style(ws_hud, pr, 2, False, bg="1a0a00", halign="center")
    style(ws_hud, pr, 3, "✓ PAP Done", bg="1a0a00", fg=DIM, size=9)
    for col in range(4, 10):
        ws_hud.cell(pr, col).fill = fill("1a0a00")
    row_height(ws_hud, pr, 26)

# ── ROW 41: Divider ───────────────────────────
divider_row(ws_hud, 41)

# ── ROW 42: BAG HEADER ────────────────────────
block_header(ws_hud, 42, "🥊  VARGA BAG WORK", bg=RED)

# ── ROWS 43-46: Bag info ──────────────────────
bag_key_b = '"P"&Settings!B3&"-D"&MID(B4,5,1)&"-BAG-1"'

style(ws_hud, 43, 1, "Target Rounds", bg=PANEL, fg=LABEL, bold=True)
ws_hud.merge_cells("B43:E43")
ws_hud.cell(43, 2).value = f"=IFERROR(INDEX(Playbook!I:I,MATCH({bag_key_b},Playbook!A:A,0)),\"—\")"
ws_hud.cell(43, 2).fill  = fill(PANEL)
ws_hud.cell(43, 2).font  = font(BODY)
ws_hud.cell(43, 2).alignment = Alignment(vertical="center")
style(ws_hud, 43, 6, "ROUNDS DONE:", bg=PANEL, fg=LABEL, bold=True, halign="right")
style(ws_hud, 43, 7, "", bg=INPUT, fg=GOLD, bold=True, halign="center")  # user input
ws_hud.merge_cells("H43:I43")
ws_hud.cell(43, 8).fill = fill(PANEL)
row_height(ws_hud, 43, 30)

style(ws_hud, 44, 1, "Combo Focus", bg=PANEL, fg=LABEL, bold=True)
ws_hud.merge_cells("B44:I44")
ws_hud.cell(44, 2).value = f"=IFERROR(INDEX(Playbook!N:N,MATCH({bag_key_b},Playbook!A:A,0)),\"—\")"
ws_hud.cell(44, 2).fill  = fill(PANEL)
ws_hud.cell(44, 2).font  = font(BODY, size=9)
ws_hud.cell(44, 2).alignment = Alignment(vertical="center", wrap_text=True)
row_height(ws_hud, 44, 45)

style(ws_hud, 45, 1, "Coach Cue", bg=PANEL, fg=LABEL, bold=True)
ws_hud.merge_cells("B45:I45")
ws_hud.cell(45, 2).value = f"=IFERROR(INDEX(Playbook!O:O,MATCH({bag_key_b},Playbook!A:A,0)),\"—\")"
ws_hud.cell(45, 2).fill  = fill(PANEL)
ws_hud.cell(45, 2).font  = font(AMBER, size=9)
ws_hud.cell(45, 2).alignment = Alignment(vertical="center", wrap_text=True)
row_height(ws_hud, 45, 36)

style(ws_hud, 46, 1, "Session Notes", bg=PANEL, fg=LABEL, bold=True)
ws_hud.merge_cells("B46:I46")
style(ws_hud, 46, 2, "", bg=INPUT, fg=WHITE)
ws_hud.cell(46, 2).alignment = Alignment(vertical="center", wrap_text=True)
row_height(ws_hud, 46, 36)

# ── ROW 47: Divider ───────────────────────────
divider_row(ws_hud, 47)

# ── ROW 48: COOLDOWN HEADER ───────────────────
block_header(ws_hud, 48, "❄️  COOLDOWN  &  STRETCH", bg=BLUE)

# ── ROW 49: Cooldown column headers ──────────
style(ws_hud, 49, 1, "Movement", bg=PANEL, fg=LABEL, bold=True, size=9)
style(ws_hud, 49, 2, "✓ Done", bg=PANEL, fg=LABEL, bold=True, size=9, halign="center")
style(ws_hud, 49, 3, "Duration", bg=PANEL, fg=LABEL, bold=True, size=9)
ws_hud.merge_cells("D49:I49")
style(ws_hud, 49, 4, "Note", bg=PANEL, fg=LABEL, bold=True, size=9)
row_height(ws_hud, 49, 22)

# ── ROWS 50-54: Cooldown slots ────────────────
for slot in range(1, 6):
    r = 49 + slot
    clr_key_b = f'"P"&Settings!B3&"-D"&MID(B4,5,1)&"-CLR-{slot}"'
    ws_hud.cell(r, 1).value = f"=IFERROR(INDEX(Playbook!G:G,MATCH({clr_key_b},Playbook!A:A,0)),\"\")"
    ws_hud.cell(r, 1).fill  = fill(PANEL)
    ws_hud.cell(r, 1).font  = font(BODY, size=10)
    ws_hud.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
    style(ws_hud, r, 2, False, bg=INPUT, halign="center")  # checkbox
    ws_hud.cell(r, 3).value = f"=IFERROR(INDEX(Playbook!I:I,MATCH({clr_key_b},Playbook!A:A,0)),\"\")"
    ws_hud.cell(r, 3).fill  = fill(PANEL)
    ws_hud.cell(r, 3).font  = font(DIM, size=9)
    ws_hud.cell(r, 3).alignment = Alignment(vertical="center")
    ws_hud.merge_cells(f"D{r}:I{r}")
    ws_hud.cell(r, 4).value = f"=IFERROR(INDEX(Playbook!J:J,MATCH({clr_key_b},Playbook!A:A,0)),\"\")"
    ws_hud.cell(r, 4).fill  = fill(PANEL)
    ws_hud.cell(r, 4).font  = font(DIM, size=9)
    ws_hud.cell(r, 4).alignment = Alignment(vertical="center", wrap_text=True)
    row_height(ws_hud, r, 28)

# ── ROW 55: Divider ───────────────────────────
divider_row(ws_hud, 55)

# ── ROW 56: COMPLETENESS HEADER ───────────────
block_header(ws_hud, 56, "📊  SESSION COMPLETENESS", bg=GREEN)
ws_hud.cell(56, 1).font = font("0a0a14", bold=True, size=11)  # dark text on green

# ── ROW 57: Completeness % ────────────────────
complete_f = (
    "=ROUND(("
    "COUNTIF(B9:B13,TRUE)+COUNTIF(B50:B54,TRUE)"   # mob + clr checkboxes (max 10)
    "+COUNTA(B18:I40)/16*8"                          # strength sets filled / max 32 cells → normalise to 8
    "+IF(ISNUMBER(G43),MIN(G43,6),0)"               # bag rounds (capped)
    ")/(10+8+6)*100,1)"
)
ws_hud.merge_cells("A57:E57")
ws_hud.cell(57, 1).value = "Completeness %"
ws_hud.cell(57, 1).fill  = fill(PANEL)
ws_hud.cell(57, 1).font  = font(LABEL, bold=True)
ws_hud.cell(57, 1).alignment = Alignment(horizontal="left", vertical="center")
ws_hud.merge_cells("F57:I57")
ws_hud.cell(57, 6).value = complete_f
ws_hud.cell(57, 6).fill  = fill(PANEL)
ws_hud.cell(57, 6).font  = font(GREEN, bold=True, size=20)
ws_hud.cell(57, 6).alignment = Alignment(horizontal="center", vertical="center")
ws_hud.cell(57, 6).number_format = '0.0"%"'
row_height(ws_hud, 57, 38)

# ── ROW 58: Hip status ────────────────────────
style(ws_hud, 58, 1, "Hip Score Today", bg=PANEL, fg=LABEL, bold=True)
ws_hud.cell(58, 2).value = "=G4"
ws_hud.cell(58, 2).fill  = fill(PANEL)
ws_hud.cell(58, 2).font  = font(WHITE, bold=True, size=14)
ws_hud.cell(58, 2).alignment = Alignment(horizontal="center", vertical="center")
ws_hud.merge_cells("C58:I58")
ws_hud.cell(58, 3).value = "=IF(G4<=2,\"🔴 HIGH ALERT — Hip mobilisation protocol is ACTIVE\",IF(G4=3,\"🟡 MODERATE — Monitor and log\",\"🟢 GOOD — Full standard protocol\"))"
ws_hud.cell(58, 3).fill  = fill(PANEL)
ws_hud.cell(58, 3).font  = font(WHITE, bold=True)
ws_hud.cell(58, 3).alignment = Alignment(vertical="center")
row_height(ws_hud, 58, 28)

# ── ROW 59: Divider ───────────────────────────
divider_row(ws_hud, 59)

# ── ROWS 60-61: Action buttons ────────────────
ws_hud.merge_cells("A60:D60")
ws_hud.cell(60, 1).value = "▶  LOG SESSION"
ws_hud.cell(60, 1).fill  = fill(RED)
ws_hud.cell(60, 1).font  = font(WHITE, bold=True, size=14)
ws_hud.cell(60, 1).alignment = Alignment(horizontal="center", vertical="center")

ws_hud.cell(60, 5).fill = fill(BG)

ws_hud.merge_cells("F60:I60")
ws_hud.cell(60, 6).value = "↺  RESET HUD"
ws_hud.cell(60, 6).fill  = fill("2a2a4a")
ws_hud.cell(60, 6).font  = font(LABEL, bold=True, size=12)
ws_hud.cell(60, 6).alignment = Alignment(horizontal="center", vertical="center")
row_height(ws_hud, 60, 44)

ws_hud.merge_cells("A61:I61")
ws_hud.cell(61, 1).value = "🔓  CHECK PHASE UNLOCK STATUS"
ws_hud.cell(61, 1).fill  = fill("1a1a2e")
ws_hud.cell(61, 1).font  = font(LABEL, size=10)
ws_hud.cell(61, 1).alignment = Alignment(horizontal="center", vertical="center")
row_height(ws_hud, 61, 28)

ws_hud.freeze_panes = "A7"  # Keep title + selector always visible

# ─────────────────────────────────────────────
# FIGHT LOG SHEET
# ─────────────────────────────────────────────
print("Building Fight Log...")

log_headers = ["Date", "Day", "Phase", "Hip_Score"]
for e in range(1, 5):
    for s in range(1, 5):
        log_headers.append(f"Ex{e}_S{s}_kg")
        log_headers.append(f"Ex{e}_S{s}_reps")
log_headers += ["Mob_Done", "Clr_Done", "Bag_Rounds", "Notes", "Completeness_%"]

for ci, h in enumerate(log_headers, 1):
    c = ws_log.cell(row=1, column=ci, value=h)
    c.fill = fill("0d2a0d")
    c.font = font(WHITE, bold=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")

col_width(ws_log, 1, 16)
col_width(ws_log, 2, 8)
col_width(ws_log, 3, 8)
col_width(ws_log, 4, 10)
for col in range(5, 38):
    col_width(ws_log, col, 9)
col_width(ws_log, 38, 8)
col_width(ws_log, 39, 8)
col_width(ws_log, 40, 12)
col_width(ws_log, 41, 30)
col_width(ws_log, 42, 16)

ws_log.freeze_panes = "A2"

# ─────────────────────────────────────────────
# STATS SHEET
# ─────────────────────────────────────────────
print("Building Stats...")

ws_stat.merge_cells("A1:D1")
ws_stat.cell(1, 1).value = "📈  PERFORMANCE STATS"
ws_stat.cell(1, 1).fill  = fill(RED)
ws_stat.cell(1, 1).font  = font(WHITE, bold=True, size=14)
ws_stat.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
ws_stat.row_dimensions[1].height = 40

stat_rows = [
    (2,  "",                      ""),
    (3,  "TOTAL SESSIONS LOGGED", "=COUNTA(FightLog!A:A)-1"),
    (4,  "CURRENT PHASE",         "=Settings!B3"),
    (5,  "PHASE UNLOCK STATUS",   "=Settings!B10"),
    (6,  "",                      ""),
    (7,  "AVG COMPLETENESS %",    "=IFERROR(AVERAGE(FightLog!AO:AO),\"-\")"),
    (8,  "BEST SESSION %",        "=IFERROR(MAX(FightLog!AO:AO),\"-\")"),
    (9,  "LOWEST HIP SCORE",      "=IFERROR(MIN(FightLog!D:D),\"-\")"),
    (10, "AVG HIP SCORE",         "=IFERROR(AVERAGE(FightLog!D:D),\"-\")"),
]

for row_n, label, val in stat_rows:
    c_label = ws_stat.cell(row_n, 1, value=label)
    c_label.fill = fill(PANEL)
    c_label.font = font(LABEL, bold=bool(label))
    c_label.alignment = Alignment(vertical="center")

    c_val = ws_stat.cell(row_n, 2, value=val if val else "")
    c_val.fill = fill(INPUT)
    c_val.font = font(GOLD, bold=True)
    c_val.alignment = Alignment(horizontal="center", vertical="center")

col_width(ws_stat, 1, 30)
col_width(ws_stat, 2, 20)

# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
out_path = os.path.join(script_dir, "fighters-os.xlsx")
wb.save(out_path)
print(f"\n✅ Built successfully: {out_path}")
print("Next steps:")
print("  1. Upload fighters-os.xlsx to Google Drive")
print("  2. Open it in Google Sheets (right-click > Open with Google Sheets)")
print("  3. Paste fighters-os-lite.gs into Extensions > Apps Script")
print("  4. Run addCheckboxesAndValidation() — takes under 30 seconds")
print("  5. Run all three button scripts: logSession / resetHUD / checkPhase")
